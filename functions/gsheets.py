""" Копирование таблицы формата XLSX в таблицу Google Sheets """
from constant import GSHEETS_TITLE
import openpyxl
import gspread


# Функция для переноса Excel-таблицы в Google Sheets.
# Table_path - путь к таблице, которую нужно обработать.
# Course - номер курса. Используется при внесении данных в листы таблицы ("Курс [course]"):
def gs_transfer(table_path: str, course: int) -> None:

    # Открытие таблицы:
    wb = openpyxl.open(table_path)
    ws = wb.active

    # Создание матрицы с данными таблицы:
    matrix = []

    # Сохраняем значение каждой ячейки таблицы в матрицу:
    for i in range(1, ws.max_row + 1):
        row = []
        for j in range(1, ws.max_column + 1):
            row.append(ws.cell(row=i, column=j).value)
        matrix.append(row)

    # Предоставляем боту доступ к Google Sheets:
    gc = gspread.service_account(filename='functions/creds.json')

    # Открываем таблицу Google Sheets:
    worktable = gc.open(GSHEETS_TITLE)
    worksheet = worktable.worksheet(f"Курс {course}")

    # Чистим устаревшую информацию в Google Sheets. Затем - обновляем:
    worksheet.clear()
    worksheet.update(matrix, 'A1')
