""" Загрузка таблицы с расписанием, разделение объединённых ячеек. Подготовка таблицы к парсингу """
from constant import *
import openpyxl


# Получение индекса границы (левого или правого) для цикла разделения объединённых ячеек:
def get_border_column_index(border: str) -> int:
    if border[1].isdigit():
        return COLUMNS.index(border[0])
    return COLUMNS.index(border[:2])


# Функция-разделитель объединённых в таблице ячеек.
# Table_path - путь к таблице, которую нужно обработать.
# Destination - путь, по которому нужно сохранить обновлённую таблицу:
def separate_merges(table_path: str, destination: str) -> None:

    # Открытие таблицы. Создание переменной WS для чтения информации:
    wb = openpyxl.open(table_path)
    ws = wb.active

    # Создаём список диапазонов объединённых ячеек:
    merged_cells = list(map(str, ws.merged_cells.ranges))

    # Разделяем каждый объединённый диапазон:
    for merge in merged_cells:
        ws.unmerge_cells(range_string=merge)

    # Форматируем данные о диапазонах для работы с ними.
    # Каждая строка формата A1:Z1 преобразуется в список ["A1", "Z1"]
    merged_cells = [d.split(":") for d in merged_cells]

    # Редактирование таблицы: разделение всех объединённых в таблице ячеек:
    for merge_cell in merged_cells:

        # Получение границ диапазона объединённых ячеек (левой и правой колонки):
        left_col, right_col = merge_cell

        # Значение, которое должно быть в каждой ячейке, относящейся к диапазону
        top_left_value = ws[left_col].value

        # Создаём левый и правый индексы границ:
        start, finish = get_border_column_index(left_col), get_border_column_index(right_col)

        # Получаем номер ряда, по которому будет пробегаться цикл:
        row = left_col[2:]
        if left_col[1].isdigit():
            row = left_col[1:]

        # Начиная от левой границы, вносим информацию в каждую ячейку ряда row, пока не дойдём до правой границы:
        for i in range(start, finish + 1):
            ws[f'{COLUMNS[i]}{row}'] = top_left_value

    # Сохранение таблицы:
    wb.save(destination)