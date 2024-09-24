""" Парсинг таблицы для автоматического обновления информации для Телеграм-бота """
from constant import *
import openpyxl


# Функция создания цвета, заполняющего ячейку таблицы:
def fill_color(code: str) -> openpyxl.styles.PatternFill:
    return openpyxl.styles.PatternFill(start_color=code, fill_type='solid')


# Форматирование строки, преобразование в лаконичный вид:
def formating_string(string: str) -> str:
    text = list()  # Список для слов
    links = list()  # Список для ссылок

    # Удаление лишних слов в строке с информацией:
    for rlt in REMOVEDLINKTEXT:
        string = string.replace(rlt, " ")
        string = string.replace(rlt.lower(), " ")

    # Удаление лишних точек:
    string = string.replace("..", ".")

    # Разделение каждого слова по пробелу:
    words = string.split()

    # Анализируем каждое отдельно взятое слово.
    # Если данное слово является ссылкой, то добавляем этот элемент в список links.
    # В противном случае добавляем элемент в список text:
    for w in words:
        if "http" in w:
            links.append(w)
        else:
            text.append(w)

    # Создаём строку: объединяем все слова, ставим между ними пробел.
    # Каждую новую ссылку выводим на отдельной строке:
    text = " ".join(text)
    links = "\n".join(links)

    # Возвращаем готовую строку:
    return f"💻 Описание пары:\n\n{text}\n\n📚 Ссылки на трансляцию / материалы:\n\n{links}\n\n==============="


# Функция-парсер.
# Table_path - путь к таблице, которую нужно обработать.
# Destination - путь, по которому нужно сохранить обновлённую таблицу.
# Course_index - индекс курса (используется при сопоставлении информации):
def parsing(table_path: str, destination: str, course_index: int) -> None:

    # Создаём новую таблицу:
    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active

    # Открываем таблицу, из которой будет получать информацию:
    wb_table = openpyxl.open(table_path)
    ws_table = wb_table.active

    # Вводим ячейки <День Недели>, <Начало>, <Окончание>.
    # Красим их в определённый цвет, устанавливаем ширину для первых трёх столбцов таблицы:
    for j in range(1, 4):
        ws_out.cell(row=1, column=j).value = DATA_COLUMNS[j - 1][0]
        ws_out.cell(row=1, column=j).fill = fill_color(DATA_COLUMNS[j - 1][1])
        ws_out.column_dimensions[COLUMNS[j - 1]].width = CWIDTH

    # Устанавливаем цвет для ячеек - названий групп.
    # Вводим названия групп в ячейки, красим их, устанавливаем ширину остальных столбцов таблицы:
    group_fill_color = fill_color(LIGHT_GRAY)
    for j in range(4, len(GROUPS[course_index]) + 4):
        ws_out.column_dimensions[COLUMNS[j - 1]].width = CWIDTH
        ws_out.cell(row=1, column=j).value = GROUPS[course_index][j - 4]
        ws_out.cell(row=1, column=j).fill = group_fill_color

    # Получаем ширину таблицы, из которой берётся информация:
    table_width = ws_table.max_column

    # Значение дня недели (1-6 - от понедельника до субботы соответственно):
    current_day = 0

    # Цвет для ячеек с информацией о паре:
    not_empty_fill_color = fill_color(BEIGE)

    # Обрабатываем информацию из таблицы в этом цикле, форматируем её и добавляем в новую таблицу.
    # Алгоритм - проходим по каждой ячейке с индексами i1, где i - номер ряда, начиная с 1, j - номер столбца, j = 1.
    # Если в данной ячейке содержится день недели - увеличиваем счётчик current_day.
    # Иначе если в данной ячейке (i1) содержится время начала и конца пары, то пробегаемся по всем
    # элементам ряда i и столбца j и вносим их в таблицу, перед этим внеся в первые три ряда
    # информацию о дне недели (current_day), начале (xx:xx) и конце (xx:xx) пары:
    current_row_index = 2
    for i in range(2, ws_table.max_row):

        # Устанавливаем высоту каждого ряда:
        ws_out.row_dimensions[i].height = CHEIGHT

        # Получаем информацию из ячейки Cell[i][1]:
        cell_i1 = ws_table.cell(row=i, column=1).value

        # Проверка на наличие строки в ячейке:
        if type(cell_i1) is str:

            # Если получен день недели - увеличиваем счётчик:
            if ':' not in cell_i1:
                current_day += 1

            # Иначе если получено время - пробегаемся по ряду i и всем столбцам этого ряда,
            # внося каждый столбец ряда i в новую таблицу:
            elif ':' in cell_i1:

                # Разделяем время начала и конца пары:
                begin, end = cell_i1.split("-")

                # На первые три столбца ряда i устанавливается
                # информация о дне недели, времени начала и времени конца пары:
                ws_out.cell(row=current_row_index, column=1).value = current_day
                ws_out.cell(row=current_row_index, column=2).value = begin
                ws_out.cell(row=current_row_index, column=3).value = end

                # Вводим в каждый столбец ряда i информацию о паре.
                # (Прим.) Создан флаг, выполняющий функцию проверки непустоты
                # хотя-бы одного столбца с информацией о паре на ряду i.
                # В противном случае данный пустой ряд не будет записан:
                flag = True
                for j in range(2, table_width + 1):
                    ij_tableValue = ws_table.cell(row=i, column=j).value

                    # Если информация о паре содержится, то удаляем лишние пробелы и красим ячейку:
                    if ij_tableValue is not None and (ij_tableValue not in ("", " ", "\n", " ")):
                        ij_tableValue = formating_string(" ".join(ij_tableValue.split()))

                        flag = False
                        ws_out.cell(row=current_row_index, column=j + 2).fill = not_empty_fill_color

                    # Заполняем ячейку:
                    ws_out.cell(row=current_row_index, column=j + 2).value = ij_tableValue

                if not flag:
                    current_row_index += 1

    # Сохранение готовой таблицы:
    wb_out.save(destination)